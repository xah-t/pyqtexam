import datetime
import xlsxwriter
import sqlite3
import openpyxl
from xlrd import open_workbook
import subprocess


fieldlist_ = []
work_cost_ = []
material_cost_ = []
material_dict = {}


def download_catalog(path):

    """Открываю xlsx файл, записываю в material_cost"""
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    connect_to_db = sqlite3.connect('fieldlist_var2.db')
    cursor_connect_to_db = connect_to_db.cursor()
    for i in range(1, sheet_obj.max_row):
        cell_obj = sheet_obj.cell(row=i, column=1)
        cell_obj_1 = sheet_obj.cell(row=i, column=2)
        values_ = (cell_obj.value, cell_obj_1.value)
        print(values_)
        sql_ = "insert into material_cost (material_mark, material_cost_rub) values(?, ?)"
        cursor_connect_to_db.execute(sql_, values_)

    #connect_to_db.commit()


def update_table_work_cost():

    """Открываю БД, выбираю из fieldlist артикулы и подстваляю в work_cost"""
    values_ = []
    connect_to_db = sqlite3.connect('fieldlist_var2.db')
    cursor_connect_to_db = connect_to_db.cursor()
    cursor_connect_to_db.execute("SELECT articul from fieldlist")
    for row in cursor_connect_to_db:
        values_.append(row)
        print(values_)

    sql_ = "insert into work_cost (detail) values (?)"
    for i in range(len(list(values_))):
        val = values_[i]
        cursor_connect_to_db.execute(sql_, val)
    # connect_to_db.commit()


"""
    cursor_material_cost_.execute("SELECT * FROM material_cost")
    for row in cursor_material_cost_:
        material_cost_.append(row)
    print(material_cost_)
    
# with connect_to_db:
#     cursor_fieldlist_ = connect_to_db.cursor()
#     cursor_fieldlist_.execute("SELECT * FROM fieldlist")  # ORDER BY id DESC LIMIT 10
#     cursor_work_cost_ = connect_to_db.cursor()
#     cursor_work_cost_.execute("SELECT time_days FROM work_cost")
#     for row in cursor_fieldlist_:
#         fieldlist_.append(row)
#     for row in cursor_work_cost_:
#         work_cost_.append(row)
#     print(fieldlist_)
#     print(work_cost_)
"""

if __name__ == "__main__":
    path = 'C:/python/VKR/pyqtexam/material_catalog.xlsx'
    download_catalog(path)

