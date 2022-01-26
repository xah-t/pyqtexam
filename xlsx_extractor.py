import datetime
import xlsxwriter
import openpyxl
import sqlite3
import os
import subprocess


class Extractor():
    """Класс для вывода Реестра в xlsx файл.

    Functions:
    init_tables,
    close,
    show,
    check_synchro_in_db.
    """

    def init_tables(self):
        """Создает xlsx файл, и добавляет в него данные из БД."""

        fieldlist_ = []
        work_cost_ = []
        material_cost_ = []
        total_cost_ = []
        connect_to_db = sqlite3.connect('fieldlist_var2.db')  # fieldlist_var2.db

        """Попробовать JOINить таблицы, чтобы резулььат в xls приходил синхронизированный"""
        #DELETE FROM work_cost WHERE (SELECT COUNT(*) FROM fieldlist WHERE fieldlist.articul = work_cost.detail) = 0

        with connect_to_db:
            cursor_fieldlist_ = connect_to_db.cursor()
            cursor_fieldlist_.execute("SELECT id, articul, name FROM fieldlist")
            for row in cursor_fieldlist_:
                fieldlist_.append(row)
            cursor_work_cost_ = connect_to_db.cursor()
            cursor_work_cost_.execute("SELECT labour, work_cost_rub FROM work_cost, fieldlist "
                                      "WHERE work_cost.detail = fieldlist.articul")
            for row in cursor_work_cost_:
                work_cost_.append(row)
            cursor_material_cost_ = connect_to_db.cursor()
            cursor_material_cost_.execute(f"SELECT material_cost.material_cost_rub * fieldlist.material_rate_kg AS mfinal_cost "
                                          f"FROM material_cost, fieldlist, work_cost WHERE material_cost.material_mark "
                                          f"LIKE fieldlist.material AND work_cost.detail = fieldlist.articul")
            for row in cursor_material_cost_:
                material_cost_.append(row)
            cursor_total_cost_ = connect_to_db.cursor()
            cursor_total_cost_.execute(f"SELECT work_cost.work_cost_rub + material_cost.material_cost_rub * fieldlist.material_rate_kg "
                                       f"AS final_cost FROM work_cost, material_cost, "
                                       f"fieldlist WHERE material_cost.material_mark LIKE fieldlist.material "
                                       f"AND work_cost.detail = fieldlist.articul")
            for row in cursor_total_cost_:
                total_cost_.append(row)

        # Create an new Excel file and add a worksheet.
        workbook = xlsxwriter.Workbook('Реестр_' + str(datetime.date.today()) + '.xlsx')
        worksheet = workbook.add_worksheet(str(datetime.date.today()))

        # Add a bold format to use to highlight cells.
        bold = workbook.add_format({'bold': True})

        # Write data headers.
        worksheet.write('A1', 'ID', bold)
        worksheet.write('B1', 'Децимальный №', bold)
        worksheet.write('C1', 'Наименование', bold)
        worksheet.write('D1', 'Трудоемкость, н/ч', bold)
        worksheet.write('E1', 'Стоимость работ, руб. с НДС', bold)
        worksheet.write('F1', 'Cтоимость материала, руб. с НДС', bold)
        worksheet.write('G1', 'Общая стоимость, руб. с НДС', bold)

        # Write some data.
        for num_row, row_data in enumerate(fieldlist_):  # expenses
            for num_col, col_data in enumerate(row_data):
                worksheet.write(num_row+1, num_col, col_data)
        for num_row, row_data in enumerate(work_cost_):
            for num_col, col_data in enumerate(row_data):
                worksheet.write(num_row+1, num_col+3, col_data)
        for num_row, row_data in enumerate(material_cost_):
            for num_col, col_data in enumerate(row_data):
                worksheet.write(num_row+1, num_col+5, col_data)
        for num_row, row_data in enumerate(total_cost_):
            for num_col, col_data in enumerate(row_data):
                worksheet.write(num_row+1, num_col+6, col_data)
        return workbook

    def close(self):
        """Закрывает xlsx файл, после внесения изменений."""

        self.workbook = self.init_tables()
        self.workbook.close()

    def show(self):
        """Открывает для просмотра xlsx файл."""

        self.path = "C:/python/VKR/pyqtexam/Реестр_" + str(datetime.date.today()) + ".xlsx"
        subprocess.Popen(self.path, shell=True)

    def check_synchro_in_db(val):
        """Возвращает True, если значение аргумента в БД.

        Argument: val
        """
        connect_to_db = sqlite3.connect('fieldlist_var2.db')
        cursor_connect_to_db = connect_to_db.cursor()
        cur = ''
        cursor_connect_to_db.execute("SELECT articul from fieldlist where articul = ?", (val,))
        try:
            cur = cursor_connect_to_db.fetchone()[0]
            if val == cur:
                return True
        except TypeError:
            return False


def download_catalog(path):
    """Функция загрузки прайса на материал в БД.

    Argument: path
    """

    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    connect_to_db = sqlite3.connect('fieldlist_var2.db')
    cursor_connect_to_db = connect_to_db.cursor()
    for i in range(1, sheet_obj.max_row):
        cell_obj = sheet_obj.cell(row=i, column=1)
        cell_obj_1 = sheet_obj.cell(row=i, column=2)
        values_ = (cell_obj.value, cell_obj_1.value)
        sql_ = "insert into material_cost (material_mark, material_cost_rub) values(?, ?)"
        cursor_connect_to_db.execute(sql_, values_)
    connect_to_db.commit()


if __name__ == '__main__':
    upload_ = Extractor()
    upload_.init_tables()
    upload_.show()
    path = 'C:/python/VKR/pyqtexam/material_catalog.xlsx'
    #download_catalog(path)

